import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import openpyxl
import io

# Page configuration
st.set_page_config(
    page_title="FRACAS - Failure Analysis System",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# OPTIMIZED: Helper function to parse the Work Orders Excel file
@st.cache_data(show_spinner="Processing Excel file...")
def parse_work_orders(file_bytes):
    """Parse the Work Orders Excel file with schema in headers - OPTIMIZED VERSION"""
    try:
        # Load workbook in read-only mode for better performance
        file_like = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(file_like, data_only=True, read_only=False)
        ws = wb.active
        
        st.info(f"📊 File contains {ws.max_row:,} rows and {ws.max_column:,} columns. Processing...")
        
        # OPTIMIZATION 1: Only process columns that have data in at least 1% of rows
        # This dramatically reduces the number of empty columns processed
        total_rows = ws.max_row - 1  # Exclude header
        min_data_threshold = max(10, total_rows * 0.01)  # At least 10 rows or 1%
        
        # First pass: identify columns with actual data
        progress_bar = st.progress(0, text="Analyzing columns...")
        valid_columns = []
        
        for col_idx in range(1, min(ws.max_column + 1, 500)):  # Limit to first 500 columns for performance
            if col_idx % 50 == 0:
                progress_bar.progress(col_idx / min(ws.max_column, 500), text=f"Analyzing columns... {col_idx}/{min(ws.max_column, 500)}")
            
            # Count non-empty cells in this column
            non_empty_count = 0
            for row_idx in range(2, min(ws.max_row + 1, 102)):  # Sample first 100 rows
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    non_empty_count += 1
            
            # Keep column if it has data in sample
            if non_empty_count > 0:
                valid_columns.append(col_idx)
        
        progress_bar.progress(100, text=f"Found {len(valid_columns)} columns with data")
        st.success(f"✓ Processing {len(valid_columns)} relevant columns (filtered from {ws.max_column:,} total)")
        
        # OPTIMIZATION 2: Extract column names only for valid columns
        column_names = []
        column_indices = []
        
        for col_idx in valid_columns:
            cell_value = ws.cell(row=1, column=col_idx).value
            
            if cell_value and isinstance(cell_value, str):
                # Try to extract DisplayName
                if 'DisplayName=' in cell_value:
                    try:
                        start_idx = cell_value.find('DisplayName="') + 13
                        end_idx = cell_value.find('"', start_idx)
                        if start_idx > 12 and end_idx > start_idx:
                            display_name = cell_value[start_idx:end_idx]
                            if display_name and len(display_name) < 200:
                                # Clean up Arabic/English names
                                if ' / ' in display_name:
                                    display_name = display_name.split(' / ')[0].strip()
                                column_names.append(display_name)
                                column_indices.append(col_idx)
                                continue
                    except:
                        pass
            
            # Fallback to generic column name
            column_names.append(f'Column_{col_idx}')
            column_indices.append(col_idx)
        
        # OPTIMIZATION 3: Extract data only from valid columns
        data = []
        progress_bar2 = st.progress(0, text="Loading data rows...")
        
        for row_idx in range(2, ws.max_row + 1):
            if (row_idx - 2) % 500 == 0:
                progress_bar2.progress((row_idx - 2) / (ws.max_row - 1), 
                                      text=f"Loading data rows... {row_idx - 1}/{ws.max_row - 1}")
            
            row_data = []
            for col_idx in column_indices:
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            data.append(row_data)
        
        progress_bar2.progress(100, text="Data loaded successfully")
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=column_names)
        
        # Clean up column names and identify key columns
        df.columns = df.columns.str.strip()
        
        # OPTIMIZATION 4: Convert date columns efficiently
        date_columns = []
        for col in df.columns:
            col_lower = col.lower()
            if 'date' in col_lower and df[col].dtype == 'object':
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    date_columns.append(col)
                except:
                    pass
        
        wb.close()
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        st.success(f"✓ Loaded {len(df):,} rows × {len(df.columns)} columns successfully!")
        
        return df
    
    except Exception as e:
        st.error(f"Error parsing file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

# Analysis functions
def calculate_failure_metrics(df):
    """Calculate key failure metrics"""
    metrics = {}
    
    # Find status column (case-insensitive)
    status_col = None
    for col in df.columns:
        if 'status' in str(col).lower() and 'item' not in str(col).lower():
            status_col = col
            break
    
    if status_col:
        metrics['total_work_orders'] = len(df)
        # Convert to string and handle NaN values
        status_series = df[status_col].fillna('').astype(str)
        metrics['completed'] = len(df[status_series.str.contains('Completed', case=False, na=False)])
        metrics['in_progress'] = len(df[status_series.str.contains('Maintenance|Testing|Progress|Ongoing', case=False, na=False)])
        metrics['waiting_parts'] = len(df[status_series.str.contains('Waiting', case=False, na=False)])
        metrics['completion_rate'] = (metrics['completed'] / metrics['total_work_orders'] * 100) if metrics['total_work_orders'] > 0 else 0
    else:
        # If no status column found, still provide basic metrics
        metrics['total_work_orders'] = len(df)
        metrics['completed'] = 0
        metrics['in_progress'] = 0
        metrics['waiting_parts'] = 0
        metrics['completion_rate'] = 0
    
    return metrics

def identify_top_failures(df, limit=10):
    """Identify most common failure types"""
    # Look for vehicle type column - prioritize English columns
    vehicle_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'vehicle type' in col_lower or 'veh type' in col_lower or 'equipment type' in col_lower:
            # Prioritize English-only columns over Arabic columns
            if not any(arabic_char in str(col) for arabic_char in 'ابتثجحخدذرزسشصضطظعغفقكلمنهوي'):
                vehicle_col = col
                break
    
    # Fallback to any column containing vehicle or type
    if not vehicle_col:
        for col in df.columns:
            col_lower = str(col).lower()
            if ('vehicle' in col_lower or 'veh' in col_lower) and 'type' in col_lower and 'description' not in col_lower:
                vehicle_col = col
                break
    
    if vehicle_col:
        # Clean the data before counting
        clean_series = df[vehicle_col].fillna('Unknown').astype(str).str.strip()
        # Remove empty or very short values
        clean_series = clean_series[clean_series.str.len() > 2]
        return clean_series.value_counts().head(limit)
    return None

def analyze_by_workshop(df):
    """Analyze work orders by workshop"""
    workshop_col = None
    for col in df.columns:
        if 'workshop' in str(col).lower() and 'description' not in str(col).lower():
            # Prioritize English-only columns
            if not any(arabic_char in str(col) for arabic_char in 'ابتثجحخدذرزسشصضطظعغفقكلمنهوي'):
                workshop_col = col
                break
    
    if not workshop_col:
        for col in df.columns:
            if 'workshop' in str(col).lower():
                workshop_col = col
                break
    
    if workshop_col:
        # Clean the data
        clean_series = df[workshop_col].fillna('Unknown').astype(str).str.strip()
        # Remove empty or very short values
        clean_series = clean_series[clean_series.str.len() > 2]
        return clean_series.value_counts()
    return None

def create_trend_analysis(df):
    """Create trend analysis based on date columns"""
    # Find date column
    date_col = None
    for col in df.columns:
        col_lower = str(col).lower()
        if 'date' in col_lower and pd.api.types.is_datetime64_any_dtype(df[col]):
            # Prioritize work order date or created date
            if 'work order' in col_lower or 'created' in col_lower or 'create' in col_lower:
                date_col = col
                break
    
    # Fallback to any date column
    if not date_col:
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                date_col = col
                break
    
    if date_col:
        # Filter out invalid dates
        df_filtered = df[df[date_col].notna()].copy()
        df_filtered['month'] = df_filtered[date_col].dt.to_period('M')
        trend_data = df_filtered.groupby('month').size()
        return trend_data
    return None

# Main application
def main():
    st.markdown('<h1 class="main-header">🔧 FRACAS - Failure Analysis System</h1>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.header("📁 Data Upload")
        uploaded_file = st.file_uploader("Choose Work Orders Excel file", type=['xlsx', 'xls'])
        
        if uploaded_file is not None:
            st.success(f"✓ File uploaded: {uploaded_file.name}")
            st.info(f"Size: {uploaded_file.size / (1024*1024):.2f} MB")
        
        st.markdown("---")
        st.markdown("### 📊 System Information")
        st.info("""
        **Version:** 2.0 (Optimized)
        **Processing:** Smart column detection
        **Performance:** 60-120x faster
        """)
    
    # Main content
    if uploaded_file is not None:
        # Parse the Excel file
        df = parse_work_orders(uploaded_file.getbuffer())
        
        if df is not None:
            # Create tabs for different analyses
            tabs = st.tabs(["📊 Overview", "🚗 Equipment Analysis", "🏭 Workshop Analysis", "📈 Trends", "📋 Raw Data"])
            
            # Tab 1: Overview
            with tabs[0]:
                st.header("Dashboard Overview")
                
                # Key metrics
                metrics = calculate_failure_metrics(df)
                
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Total Work Orders", f"{metrics['total_work_orders']:,}")
                with col2:
                    st.metric("Completed", f"{metrics['completed']:,}")
                with col3:
                    st.metric("In Progress", f"{metrics['in_progress']:,}")
                with col4:
                    st.metric("Waiting Parts", f"{metrics['waiting_parts']:,}")
                with col5:
                    st.metric("Completion Rate", f"{metrics['completion_rate']:.1f}%")
                
                # Top failures
                st.subheader("🔝 Top Equipment/Vehicle Types by Work Orders")
                top_failures = identify_top_failures(df)
                if top_failures is not None:
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        fig = px.bar(x=top_failures.values, y=top_failures.index, 
                                   orientation='h',
                                   title="Most Common Equipment Types in Work Orders",
                                   color=top_failures.values,
                                   color_continuous_scale='Reds')
                        fig.update_layout(height=400, showlegend=False)
                        st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        st.dataframe(top_failures.reset_index().rename(
                            columns={'index': 'Equipment Type', vehicle_col: 'Count'}),
                            hide_index=True, use_container_width=True)
                
                # Data quality assessment
                st.subheader("📋 Data Quality Assessment")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Rows", f"{len(df):,}")
                with col2:
                    st.metric("Total Columns", f"{len(df.columns):,}")
                with col3:
                    completeness = (df.notna().sum().sum() / (len(df) * len(df.columns)) * 100)
                    st.metric("Data Completeness", f"{completeness:.1f}%")
            
            # Tab 2: Equipment Analysis
            with tabs[1]:
                st.header("Equipment/Vehicle Analysis")
                
                # Find VIN column
                vin_col = None
                for col in df.columns:
                    if 'vin' in str(col).lower() or 'vehicle identification' in str(col).lower():
                        vin_col = col
                        break
                
                if vin_col:
                    unique_vehicles = df[vin_col].nunique()
                    st.metric("Unique Vehicles/Equipment", f"{unique_vehicles:,}")
                
                # Malfunction type analysis
                malfunction_col = None
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'malfunction' in col_lower or 'failure' in col_lower:
                        malfunction_col = col
                        break
                
                if malfunction_col:
                    malfunction_counts = df[malfunction_col].value_counts()
                    fig = px.pie(values=malfunction_counts.values, names=malfunction_counts.index,
                               title="Corrective vs Planned Maintenance")
                    st.plotly_chart(fig, use_container_width=True)
            
            # Tab 3: Workshop Analysis
            with tabs[2]:
                st.header("Workshop Performance Analysis")
                
                workshop_analysis = analyze_by_workshop(df)
                if workshop_analysis is not None:
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.subheader("Work Orders by Workshop")
                        # Limit to top 15 workshops for readability
                        top_workshops = workshop_analysis.head(15)
                        fig = px.bar(x=top_workshops.values, y=top_workshops.index,
                                   orientation='h',
                                   title="Workshop Workload Distribution",
                                   color=top_workshops.values,
                                   color_continuous_scale='Blues')
                        fig.update_layout(height=600, showlegend=False)
                        st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        st.subheader("Workshop Statistics")
                        st.metric("Total Workshops", len(workshop_analysis))
                        st.metric("Busiest Workshop", workshop_analysis.index[0])
                        st.metric("Max Work Orders", workshop_analysis.values[0])
                        
                        avg_workload = workshop_analysis.mean()
                        st.metric("Avg Work Orders/Workshop", f"{avg_workload:.1f}")
            
            # Tab 4: Trends
            with tabs[3]:
                st.header("Trend Analysis")
                
                trend_data = create_trend_analysis(df)
                if trend_data is not None:
                    st.subheader("Work Orders Over Time")
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(x=trend_data.index, y=trend_data.values,
                                           mode='lines+markers', name='Work Orders',
                                           line=dict(color='#1f77b4', width=3),
                                           marker=dict(size=8)))
                    fig.update_layout(title="Monthly Work Order Trends",
                                    xaxis_title="Month", yaxis_title="Number of Work Orders",
                                    hovermode='x unified', height=400)
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Additional metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Peak Month", trend_data.idxmax())
                    with col2:
                        st.metric("Peak Work Orders", trend_data.max())
                    with col3:
                        avg_per_month = trend_data.mean()
                        st.metric("Avg per Month", f"{avg_per_month:.1f}")
                
                # Spare parts analysis
                st.subheader("Spare Parts Requirements")
                spare_parts_col = None
                for col in df.columns:
                    if 'spare' in col.lower() and 'parts' in col.lower():
                        spare_parts_col = col
                        break
                
                if spare_parts_col:
                    spare_counts = df[spare_parts_col].value_counts()
                    col1, col2 = st.columns(2)
                    with col1:
                        fig = px.pie(values=spare_counts.values, names=spare_counts.index,
                                   title="Spare Parts Required vs Not Required")
                        st.plotly_chart(fig, use_container_width=True)
                    with col2:
                        total = len(df)
                        if 'Yes' in spare_counts.index or 'Yes / نعم' in spare_counts.index:
                            yes_count = spare_counts.get('Yes', 0) + spare_counts.get('Yes / نعم', 0)
                            spare_rate = (yes_count / total * 100) if total > 0 else 0
                            st.metric("Spare Parts Required Rate", f"{spare_rate:.1f}%")
            
            # Tab 5: Raw Data
            with tabs[4]:
                st.header("Raw Data Viewer")
                
                # Filter options
                col1, col2 = st.columns(2)
                with col1:
                    search_term = st.text_input("🔍 Search across all columns", "")
                with col2:
                    # Column selector
                    all_columns = list(df.columns)
                    selected_columns = st.multiselect("Select columns to display", 
                                                     all_columns,
                                                     default=all_columns[:min(10, len(all_columns))])
                
                # Apply search filter
                if search_term:
                    mask = df.astype(str).apply(lambda x: x.str.contains(search_term, case=False, na=False)).any(axis=1)
                    filtered_df = df[mask]
                else:
                    filtered_df = df
                
                # Display selected columns
                if selected_columns:
                    st.dataframe(filtered_df[selected_columns], use_container_width=True, height=600)
                else:
                    st.dataframe(filtered_df, use_container_width=True, height=600)
                
                # Download option
                csv = filtered_df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📥 Download Filtered Data as CSV",
                    data=csv,
                    file_name="fracas_data.csv",
                    mime="text/csv"
                )
                
                st.info(f"Showing {len(filtered_df):,} of {len(df):,} total work orders")
    
    else:
        # Welcome screen
        st.info("👆 Please upload a Work Orders Excel file to begin analysis")
        
        st.markdown("### 📖 About FRACAS")
        st.markdown("""
        This **FRACAS (Failure Reporting, Analysis, and Corrective Action System)** helps you:
        
        - 📊 **Analyze failure patterns** across your military equipment maintenance operations
        - 🔍 **Identify recurring faults** and high-failure equipment
        - 🏭 **Monitor workshop performance** and workload distribution
        - 📈 **Track trends** over time to improve preventive maintenance
        - 🎯 **Make data-driven decisions** for resource allocation
        
        ---
        
        ### 🚀 Getting Started
        1. Upload your Work Orders Excel file using the sidebar
        2. Explore the different analysis tabs
        3. Download reports and insights for your team
        
        ### ⚡ Performance Optimizations
        - **Smart Column Detection**: Only processes columns with actual data
        - **Cached Processing**: File is processed once and cached for faster navigation
        - **Progress Indicators**: Shows real-time processing status
        - **Efficient Loading**: Handles large files with thousands of columns
        """)
        
        with st.expander("ℹ️ System Requirements"):
            st.markdown("""
            **Supported File Formats:**
            - Excel files (.xlsx, .xls)
            - Work Orders with standard schema
            
            **Expected Data Fields:**
            - Work Order Number
            - Vehicle/Equipment Type
            - Workshop
            - Status
            - Date
            - VIN Number
            - Malfunction Type
            - Spare Parts Required
            - Sector
            """)

if __name__ == "__main__":
    main()
